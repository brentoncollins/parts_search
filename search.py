from tkinter import *
from tkinter import messagebox
from xlrd import open_workbook
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side


def readxls():
	book = open_workbook("Parts_List.xlsx")
	# Open workbook
	sheet_names = book.sheet_names()
	# Get sheet names
	no_sheets = book.nsheets
	# Get number of sheets
	workbook = []
	# Create empty list for all data in workbook
	sheet_num = 0
	# Set sheet number to x, add 1 when every sheet has finished.
	for x in range(0, no_sheets):
		# Assign each sheet in every loop
		sheet = book.sheet_by_index(x)
		# Clear the item list after every sheet, this wil differentiate between the lockers.
		item_list = []

		draw = 0
		# Start at draw 0, when finding the first draw in the spreadsheet add one.
		for row in range(2, sheet.nrows):
			if "Draw" == str(sheet.cell(row, 0).value.split(" ")[0]):

				# If found draw, move to the next draw.
				draw += 1
			else:

				if "Item" in str(sheet.cell(row, 0).value.split(" ")[0]) or sheet.cell(row, 0).value == "":
					# Ignore all rows that start with item or is empty.
					pass

				else:
					# Append all items to a dict with the item as the key, all aother data are the values.

					# For the numbers that are converted into floats change them to int
					if type(sheet.cell(row, 3).value) == float:
						stock_no = int(sheet.cell(row, 3).value)
					else:
						stock_no = ""

					if type(sheet.cell(row, 2).value) == float:
						part_number = int(sheet.cell(row, 2).value)
					else:
						part_number = sheet.cell(row, 2).value

					if type(sheet.cell(row, 4).value) == float:
						part_type = int(sheet.cell(row, 4).value)
					else:
						part_type = sheet.cell(row, 4).value
					# Then append the values
					item_list.append(
								[{sheet.cell(row, 0).value: ["Draw " + str(draw),
									sheet.cell(row, 1).value,
									part_number,
									stock_no,
									part_type,
									sheet.cell(row, 5).value,
									sheet_names[sheet_num]
									]}])
		# When done with the sheet all the list of items to the workbook list.
		workbook.append(item_list)
		# Move to the next sheet.
		sheet_num += 1
	# Return all data for all items.
	return workbook


	# First create application class
class Application(object):

	def __init__(self, master):
		self.master = master

		self.read_xl()

	# Create main GUI window

		self.draws = [
			'Draw One', 'Draw Two',
			'Draw Three', 'Draw Four',
			'Draw Five', 'Draw Six',
			'Draw Seven', 'Draw Eight',
			'Draw Nine']
		self.lockers = [
					'Locker A', 'Locker B',
					'Locker C', 'Locker D',
					'Locker E', 'Locker F',
					'Locker G', 'Locker H']
		# Search Vars

		self.results = StringVar()
		self.search_var = StringVar()

		# Entry Vars

		self.item_var = StringVar()
		self.brand_var = StringVar()
		self.part_var = StringVar()
		self.stock_var = IntVar()
		self.type_var = StringVar()
		self.desc_var = StringVar()
		self.draw_var = StringVar()
		self.locker_var = StringVar()
		self.question_var = StringVar()

		self.search_var.trace("w", self.update_list)
		self.entry = Entry(master, textvariable=self.search_var, width=20)
		self.list_box_item = Listbox(master, width=25, height=25)
		self.list_box_brand = Listbox(master, width=20, height=25)
		self.list_box_part = Listbox(master, width=20, height=25)
		self.list_box_stock = Listbox(master, width=15, height=25)
		self.list_box_type = Listbox(master, width=15, height=25)
		self.list_box_disc = Listbox(master, width=50, height=25)
		self.list_box_draw = Listbox(master, width=15, height=25)
		self.list_box_locker = Listbox(master, width=15, height=25)
		self.add_entry_button = Button(master, text="Add Entry", command=self.show_entry)

		self.top_search_label = Label(
			master, text="Marandoo Fixed Plant Electrical Parts Search", font=("Helvetica", 20), anchor="n")

		self.search_label = Label(master, text="Search", font=("Helvetica", 20), fg ="red")
		self.list_box_item_label = Label(master, text="Item")
		self.list_box_brand_label = Label(master, text="Brand")
		self.list_box_part_label = Label(master, text="Part Number")
		self.list_box_stock_label = Label(master, text="Stock Number")
		self.list_box_type_label = Label(master, text="Type")
		self.list_box_disc_label = Label(master, text="Description")
		self.list_box_draw_label = Label(master, text="Draw")
		self.list_box_locker_label = Label(master, text="Locker")
		self.results_label = Label(master, textvariable=self.results, font=("Helvetica", 14), fg="blue")

		self.results_label.grid(row=1, column=2, padx=10, pady=3, columnspan=3)
		self.add_entry_button.grid(row=0, column=7, padx=10, pady=3)
		self.top_search_label.grid(row=0, column=2, padx=10, pady=3, columnspan=4)
		self.search_label.grid(row=1, column=0, padx=10, pady=3, sticky='E')
		self.list_box_item_label.grid(row=3, column=0, padx=10, pady=3)
		self.list_box_locker_label.grid(row=3, column=1, padx=10, pady=3)
		self.list_box_draw_label.grid(row=3, column=2, padx=10, pady=3)
		self.list_box_brand_label.grid(row=3, column=3, padx=10, pady=3)
		self.list_box_type_label.grid(row=3, column=4, padx=10, pady=3)
		self.list_box_stock_label.grid(row=3, column=5, padx=10, pady=3)
		self.list_box_part_label.grid(row=3, column=6, padx=10, pady=3)
		self.list_box_disc_label.grid(row=3, column=7, padx=10, pady=3)

		self.entry.grid(row=1, column=1, padx=10, pady=3, columnspan=2, sticky='W')
		self.list_box_item.grid(row=4, column=0, padx=10, pady=3)
		self.list_box_locker.grid(row=4, column=1, padx=10, pady=3)
		self.list_box_draw.grid(row=4, column=2, padx=10, pady=3)
		self.list_box_brand.grid(row=4, column=3, padx=10, pady=3)
		self.list_box_part.grid(row=4, column=4, padx=10, pady=3)
		self.list_box_stock.grid(row=4, column=5, padx=10, pady=3)
		self.list_box_type.grid(row=4, column=6, padx=10, pady=3)
		self.list_box_disc.grid(row=4, column=7, padx=10, pady=3)

		# Define all entry's and option menus.
		self.item_entry = Entry(self.master, textvariable=self.item_var, width=25)
		self.brand_entry = Entry(self.master, textvariable=self.brand_var, width=20)
		self.part_entry = Entry(self.master, textvariable=self.part_var, width=20)
		self.stock_entry = Entry(self.master, textvariable=self.stock_var, width=15)
		self.type_entry = Entry(self.master, textvariable=self.type_var, width=15)
		self.disc_entry = Entry(self.master, textvariable=self.desc_var, width=50)
		self.draw_entry_option = OptionMenu(self.master, self.draw_var, *self.draws)
		self.locker_entry_option = OptionMenu(self.master, self.locker_var, *self.lockers)
		self.draw_entry_option.config(width=9)
		self.locker_entry_option.config(width=9)

		# Set all labels
		self.search_label = Label(self.master, text="Enter New Item", font=("Helvetica", 16), fg="blue")
		self.information_label = Label(
			self.master,
			text="Fill all relevant forms then click 'Add Item'",
			font=("Helvetica", 12), fg="black")
		self.item_label_entry = Label(self.master, text="Item")
		self.brand_label_entry = Label(self.master, text="Brand")
		self.part_label_entry = Label(self.master, text="Part Number")
		self.stock_label_entry = Label(self.master, text="Stock Number")
		self.type_label_entry = Label(self.master, text="Type")
		self.disc_label_entry = Label(self.master, text="Description")
		self.draw_label_entry = Label(self.master, text="Draw")
		self.locker_label_entry = Label(self.master, text="Locker")
		self.add_button = Button(self.master, text="Add Item", command=self.show_btn_lbl)
		self.yes_button = Button(self.master, text="Add", command=self.remove_btn_lbl, bg="green")
		self.no_button = Button(self.master, text="Change", command=self.change_entry, bg="red")
		self.question_label = Label(
			self.master,
			textvariable=self.question_var,
			font=("Helvetica", 12), fg="black")

		self.item_label_entry.grid(row=7, column=0, padx=10, pady=3)
		self.locker_label_entry.grid(row=7, column=1, padx=10, pady=3)
		self.draw_label_entry.grid(row=7, column=2, padx=10, pady=3)
		self.brand_label_entry.grid(row=7, column=3, padx=10, pady=3)
		self.type_label_entry.grid(row=7, column=4, padx=10, pady=3)
		self.stock_label_entry.grid(row=7, column=5, padx=10, pady=3)
		self.part_label_entry.grid(row=7, column=6, padx=10, pady=3)
		self.disc_label_entry.grid(row=7, column=7, padx=10, pady=3)

		self.search_label.grid(row=5, column=0, padx=10, pady=3, columnspan=2)
		self.information_label.grid(row=6, column=0, padx=10, pady=3, columnspan=2)
		self.item_entry.grid(row=8, column=0, padx=10, pady=3)
		self.locker_entry_option.grid(row=8, column=1, padx=10, pady=3)
		self.draw_entry_option.grid(row=8, column=2, padx=10, pady=3)
		self.brand_entry.grid(row=8, column=3, padx=10, pady=3)
		self.part_entry.grid(row=8, column=4, padx=10, pady=3)
		self.stock_entry.grid(row=8, column=5, padx=10, pady=3)
		self.type_entry.grid(row=8, column=6, padx=10, pady=3)
		self.disc_entry.grid(row=8, column=7, padx=10, pady=3)

		self.add_button.grid(row=5, column=7, padx=10, pady=3, columnspan=2)
		self.question_label.grid(row=5, column=2, padx=10, pady=3, columnspan=5)
		self.yes_button.grid(row=6, column=6, padx=10, pady=3, sticky='E')
		self.no_button.grid(row=6, column=7, padx=10, pady=3, sticky='W')

		# Function for updating the list/doing the search.
		# It needs to be called here to populate the listbox.
		self.update_list()
		self.hide_entry()

	def update_list(self, *args):
		self.read_xl()
		self.results_label = Label(self.master, font=("Helvetica", 20), fg="red")

		search_term = self.search_var.get()

		self.list_box_item.delete(0, END)
		self.list_box_brand.delete(0, END)
		self.list_box_part.delete(0, END)
		self.list_box_stock.delete(0, END)
		self.list_box_type.delete(0, END)
		self.list_box_disc.delete(0, END)
		self.list_box_draw.delete(0, END)
		self.list_box_locker.delete(0, END)
		for y in self.items:
			for item in y:
				for key, value in item[0].items():
					if search_term.lower() in str(value).lower() or search_term.lower() in key.lower():
						self.list_box_item.insert(END, key)
						self.list_box_draw.insert(END, value[0])
						self.list_box_brand.insert(END, value[1])
						self.list_box_type.insert(END, value[2])
						self.list_box_stock.insert(END, value[3])
						self.list_box_part.insert(END, value[4])
						self.list_box_disc.insert(END, value[5])
						self.list_box_locker.insert(END, value[6])

		for x in range(0, self.list_box_item.size()):
			if x & 1:
				self.list_box_locker.itemconfig(x, {'bg': 'light blue'})
				self.list_box_draw.itemconfig(x, {'bg': 'light blue'})
				self.list_box_disc.itemconfig(x, {'bg': 'light blue'})
				self.list_box_type.itemconfig(x, {'bg': 'light blue'})
				self.list_box_stock.itemconfig(x, {'bg': 'light blue'})
				self.list_box_part.itemconfig(x, {'bg': 'light blue'})
				self.list_box_brand.itemconfig(x, {'bg': 'light blue'})
				self.list_box_item.itemconfig(x, {'bg': 'light blue'})
		self.results.set("{} Results Found".format(int(self.list_box_item.size())))

	def hide_entry(self):
		self.search_label.grid_remove()
		self.item_entry.grid_remove()
		self.locker_entry_option.grid_remove()
		self.draw_entry_option.grid_remove()
		self.brand_entry.grid_remove()
		self.part_entry.grid_remove()
		self.stock_entry.grid_remove()
		self.type_entry.grid_remove()
		self.disc_entry.grid_remove()

		self.add_button.grid_remove()
		self.question_label.grid_remove()
		self.yes_button.grid_remove()
		self.no_button.grid_remove()

		self.item_label_entry.grid_remove()
		self.locker_label_entry.grid_remove()
		self.draw_label_entry.grid_remove()
		self.brand_label_entry.grid_remove()
		self.type_label_entry.grid_remove()
		self.stock_label_entry.grid_remove()
		self.part_label_entry.grid_remove()
		self.disc_label_entry.grid_remove()
		self.information_label.grid_remove()

	def show_entry(self):
		self.search_label.grid()
		self.item_entry.grid()
		self.locker_entry_option.grid()
		self.draw_entry_option.grid()
		self.brand_entry.grid()
		self.part_entry.grid()
		self.stock_entry.grid()
		self.type_entry.grid()
		self.disc_entry.grid()

		self.item_label_entry.grid()
		self.locker_label_entry.grid()
		self.draw_label_entry.grid()
		self.brand_label_entry.grid()
		self.type_label_entry.grid()
		self.stock_label_entry.grid()
		self.part_label_entry.grid()
		self.disc_label_entry.grid()
		self.information_label.grid()
		self.add_button.grid()

	def show_btn_lbl(self):

		# Hide enter new entry until option chosen.
		self.add_button.grid_remove()
		# Run function so show labels and buttons when adding item.
		self.question_label.grid()
		self.no_button.grid()
		self.yes_button.grid()
		try:

			self.question_var.set("Is this information correct?")
			# "Are you sure you want to add\n"
			# 	" {} - {} - {} - {} - {} - {} - {} - {}?".format(
			# 		self.item_var.get().title(),
			# 		self.draw_var.get(),
			# 		self.locker_var.get(),
			# 		self.brand_var.get().title(),
			# 		self.part_var.get().upper(),
			self.stock_var.get()
			# 		self.type_var.get().upper(),
			# 		self.desc_var.get().title()))
		except TclError:
			messagebox.showerror("Error", "Please only use an integer for the Stock Number.")

	def change_entry(self):
		self.question_label.grid_remove()
		self.no_button.grid_remove()
		self.yes_button.grid_remove()
		self.add_button.grid()

	def remove_btn_lbl(self):
		# Hide again when trying again.
		self.question_label.grid_remove()
		self.no_button.grid_remove()
		self.yes_button.grid_remove()
		# Show add item button again.
		self.add_button.grid()
		self.add_item()

	def read_xl(self):
		try:
			self.items = readxls()
		except FileNotFoundError:
			messagebox.showerror(
				"File not found",
				"Please ensure that 'Parts_List.xlsx' is located in the same directory as the executable."
				"\n\nNo results will be shown.")

			self.items = []

	def add_item(self):
		# Open workbook to view all items with xlrd to find the correct input position.
		book = open_workbook("Parts_List.xlsx")

		# Open the workbook to add a row with openpyxl
		wb = load_workbook(filename='Parts_List.xlsx')

		# Get the sheet names (xlrd)
		sheet_names = book.sheet_names()

		# Get the number of sheets (xlrd)
		no_sheets = book.nsheets

		# Loop threw all of the sheets to find the same sheet name as the locker variable.
		for x in range(0, no_sheets):
			if sheet_names[x] == self.locker_var.get():
				# When found define the sheet we need to work with.
				sheet = book.sheet_by_index(x)

				# Loop threw all rows on sheet o find the correct draw number.
				for row in range(2, sheet.nrows):
					# When found open the worksheet in openpyxl
					if self.draw_var.get() in str(sheet.cell(row, 0)):
						ws = wb.worksheets[x]
						# Insert row in the row number found plus 3, this will put it at the position just below the
						# header for the draw.
						ws.insert_rows(row + 3)
						# Define the border so that it does not copy the bold from header.

						thin_border = Border(
												right=Side(style='medium'),
												bottom=Side(style=None),
												top=Side(style='medium')
												)

						try:
							self.stock_var.get()

						except TclError:

							messagebox.showerror("Error", "Please only use an integer for the Stock Number.")

						# Set all the variables and borders into the cells.
						ws.cell(column=1, row=row + 3, value=self.item_var.get().title()).border = thin_border
						ws.cell(column=1, row=row + 4).border = Border(top=None, right=Side(style='medium'))
						ws.cell(column=2, row=row + 3, value=self.brand_var.get().title()).border = thin_border
						ws.cell(column=2, row=row + 4).border = Border(top=None, right=Side(style='medium'))
						ws.cell(column=3, row=row + 3, value=self.type_var.get().upper()).border = thin_border
						ws.cell(column=3, row=row + 4).border = Border(top=None, right=Side(style='medium'))
						ws.cell(column=4, row=row + 3, value=self.stock_var.get()).border = thin_border
						ws.cell(column=4, row=row + 4).border = Border(top=None, right=Side(style='medium'))
						ws.cell(column=5, row=row + 3, value=self.part_var.get().upper()).border = thin_border
						ws.cell(column=5, row=row + 4).border = Border(top=None, right=Side(style='medium'))
						ws.cell(column=6, row=row + 3, value=self.desc_var.get().title()).border = thin_border
						ws.cell(column=6, row=row + 4).border = Border(top=None, right=Side(style='medium'))

						# Ensure workbook not open or permission available.
						try:
							wb.save("Parts_List.xlsx")
						except PermissionError:
							messagebox.showerror(
								"Error",
								"No permission to write to the Parts List, please close the "
								"document or gain permission to write to the file.")
							break

						# Remove all entry input. ( Maybe add a clear button on main page instead)
						self.item_entry.delete(0, 'end')
						self.brand_entry.delete(0, 'end')
						self.part_entry.delete(0, 'end')
						self.stock_var.set(0)
						self.type_entry.delete(0, 'end')
						self.disc_entry.delete(0, 'end')
					# Try update main list here.
						# Remove question buttons, set question label to done.
						print("done")
						self.question_label.grid()
						self.question_var.set("Done")

						self.read_xl()
						self.update_list()


root = Tk()
root.title('Electrical workshop parts search.')

app = Application(root)
print('Starting mainloop()')

root.mainloop()

# pyinstaller --onefile --hidden-import=tkinter.messagebox -i=coal.ico -w --windowed --noconsole search.py

